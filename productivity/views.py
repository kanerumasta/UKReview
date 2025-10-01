from django.shortcuts import render
from django.http import HttpResponse
from django.core.paginator import Paginator
import random
import csv
from datetime import datetime, timedelta, date
from jobs.models import ProvisionJob
from django.contrib.auth import get_user_model
User = get_user_model()
from django.views.decorators.cache import cache_page


import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.db.models import Sum, Count, F, ExpressionWrapper, DurationField, Q, FloatField, Func, Case, When, Value, FloatField
from django.contrib.auth import get_user_model
from datetime import timedelta
from django.db.models.functions import Cast
from settings.models import JobSettings

class ExtractEpoch(Func):
    function = "EXTRACT"
    template = "%(function)s(EPOCH FROM %(expressions)s)"
    output_field = FloatField()



def get_user_productivity():
    settings = JobSettings.objects.first()

    users = (
        User.objects.annotate(
            # total jobs completed
            total_jobs_completed=Count(
                "jobs", filter=Q(jobs__status="completed"), distinct=True
            ),
            total_jobs_assigned=Count("jobs", distinct=True),
            total_enactment_allocated=Count("enactment_assignments", distinct=True),
            total_time_spent=Sum(
                ExpressionWrapper(
                    F("jobs__sessions__ended_at") - F("jobs__sessions__started_at"),
                    output_field=DurationField(),
                )
            ),
        )
        .annotate(total_seconds=ExtractEpoch(F("total_time_spent")))
        .annotate(
            total_hours=ExpressionWrapper(
                F("total_seconds") / 3600.0,
                output_field=FloatField(),
            ),
            average_jobs_per_hour=ExpressionWrapper(
                F("total_jobs_completed") / (F("total_seconds") / 3600.0),
                output_field=FloatField(),
            ),
            # 👇 pick quota based on is_part_time
            # effective_quota=Case(
            #     When(is_part_time=True, then=Value(settings.parttime_quota)),
            #     default=Value(settings.quota),
            #     output_field=FloatField(),
            # ),
            
            productivity_ratio=ExpressionWrapper(
                # F("average_jobs_per_hour") / F("effective_quota") * 100,
                F("average_jobs_per_hour") / settings.quota * 100,
                output_field=FloatField(),
            ),
        )
    ).order_by('productivity_ratio')
    return users

# @cache_page(60 * 5) # Cache the response for 5 minutes
def index(request):

    users = get_user_productivity()

    context = {
        "active_page": "productivity",
        "users": users,
    }

    return render(request, "productivity/index.html", context)

def export_to_excel(request):
    # Get user productivity data
    users = get_user_productivity()

    # Create an in-memory workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productivity Report"

    # Define the headers
    headers = [
        "Username",
        "Total Jobs Assigned",
        "Total Jobs Completed",
        "Total Enactments Allocated",
        "Total Time Spent (hours)",
        "Average Jobs per Hour",
        "Effective Quota",
        "Productivity Ratio (%)",
    ]

    # Add headers to the first row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Add data for each user
    for row_num, user in enumerate(users, start=2):
        ws.cell(row=row_num, column=1, value=user.username)
        ws.cell(row=row_num, column=2, value=user.total_jobs_assigned)
        ws.cell(row=row_num, column=3, value=user.total_jobs_completed)
        ws.cell(row=row_num, column=4, value=user.total_enactment_allocated)
        hours_cell = ws.cell(row=row_num, column=5, value=user.total_hours)
        hours_cell.number_format = "0.00"

        avg_jobs_cell = ws.cell(row=row_num, column=6, value=user.average_jobs_per_hour)
        avg_jobs_cell.number_format = "0.00"

        quota_cell = ws.cell(row=row_num, column=7, value=user.effective_quota)
        quota_cell.number_format = "0.00"

        productivity_cell = ws.cell(row=row_num, column=8, value=user.productivity_ratio)
        productivity_cell.number_format = "0.00"
    # Adjust column width to fit data
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in range(1, len(users) + 2):  # Include header row
            cell = ws[column + str(row)]
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    date_time = datetime.now().strftime("%m%d%Y_%H%M%S") 

    # Create an HTTP response with the Excel file as an attachment
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="user_productivity_report_{date_time}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response



# def export_to_excel(request):
    # try:
    #     # --- Reuse your filters ---
    #     jobs = ProvisionJob.objects.select_related(
    #         "user", "provision", "enactment_assignment__enactment"
    #     ).prefetch_related("sessions").order_by('-last_edited')

    #     start_date = request.GET.get("start_date")
    #     end_date = request.GET.get("end_date")
    #     user_id = request.GET.get("user_id")

    #     if user_id:
    #         try:
    #             jobs = jobs.filter(user__id=int(user_id))
    #         except ValueError:
    #             pass

    #     if start_date:
    #         try:
    #             start = datetime.strptime(start_date, "%Y-%m-%d").date()
    #             jobs = jobs.filter(date_assigned__date__gte=start)
    #         except Exception as e:
    #             print("Start date filter error:", e)

    #     if end_date:
    #         try:
    #             end = datetime.strptime(end_date, "%Y-%m-%d").date()
    #             jobs = jobs.filter(date_assigned__date__lte=end)
    #         except Exception as e:
    #             print("End date filter error:", e)




    #     # --- Create workbook ---
    #     wb = openpyxl.Workbook()
    #     ws = wb.active
    #     ws.title = "User Productivity Report"

    #     # --- Header row ---
    #     headers = [
    #         'Date Assigned',
    #         'User ID',
    #         'User Name',
    #         'Enactment',
    #         'Provision Ref(s)',
    #         'Start Date',
    #         'End Date',
    #         'Time Spent (hrs)',
    #         'Efficiency (%)'
    #     ]
    #     header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    #     for col_num, header in enumerate(headers, 1):
    #         cell = ws.cell(row=1, column=col_num, value=header)
    #         cell.font = Font(bold=True, color="000000")
    #         cell.fill = header_fill

    #     # --- Data rows ---
    #     row_num = 2
    #     for job in jobs:
    #         try:
    #             time_spent = round(job.total_time_minutes / 60, 2) if job.total_time_minutes else 0
    #             hourly_quota = 50
    #             output = 1 if job.status == "completed" else 0
    #             efficiency = round((output / (hourly_quota * time_spent)) * 100, 2) if time_spent > 0 else 0

    #             if job.enactment_assignment and job.enactment_assignment.enactment:
    #                 enactment_title = job.enactment_assignment.enactment.title
    #             elif job.provision and hasattr(job.provision, "enactment") and job.provision.enactment:
    #                 enactment_title = job.provision.enactment.title
    #             else:
    #                 enactment_title = None

    #             ws.cell(row=row_num, column=1, value=job.date_assigned.strftime("%m/%d/%Y") if job.date_assigned else "")
    #             ws.cell(row=row_num, column=2, value=job.user.username if job.user else "")
    #             ws.cell(row=row_num, column=3, value=job.user.get_full_name() if job.user else "")
    #             ws.cell(row=row_num, column=4, value=enactment_title or "")
    #             ws.cell(row=row_num, column=5, value=job.provision.title if job.provision else "")
    #             ws.cell(row=row_num, column=6, value=job.start_date.strftime("%m/%d/%Y %I:%M %p") if job.start_date else "")
    #             ws.cell(row=row_num, column=7, value=job.end_date.strftime("%m/%d/%Y %I:%M %p") if job.end_date else "")
    #             ws.cell(row=row_num, column=8, value=time_spent)
    #             ws.cell(row=row_num, column=9, value=f"{efficiency}%")

    #             row_num += 1
    #         except Exception as e:
    #             print(f"Error exporting job {job.id}: {e}")
    #             continue

    #     # --- Auto-fit columns ---
    #     for col_num, _ in enumerate(headers, 1):
    #         column_letter = get_column_letter(col_num)
    #         ws.column_dimensions[column_letter].auto_size = True

    #     # --- Dynamic filename ---
    #     now = datetime.now().strftime("%Y%m%d_%H%M%S")
    #     filename = f"user_productivity_report_{now}.xlsx"

    #     response = HttpResponse(
    #         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #     )
    #     response["Content-Disposition"] = f'attachment; filename="{filename}"'
    #     wb.save(response)
    #     return response

    # except Exception as e:
    #     print("Export to Excel error:", e)
    #     return HttpResponse("Error generating Excel file", status=500)