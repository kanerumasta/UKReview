from django.shortcuts import render, get_object_or_404
from enactments.models import Batch, Enactment
from jobs.models import ProvisionJob
from defects.models import DefectLog, DefectCategory, DefectOption
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Count, Sum, Exists, OuterRef, Value, CharField, Case, When
from openpyxl.utils import get_column_letter, range_boundaries
from django.utils import timezone 
from django.core.paginator import Paginator
from django.db.models import Q
import os
from django.conf import settings
from openpyxl import load_workbook
import io
from copy import copy
from django.core.files import File
import tempfile
from datetime import datetime
from .models import ReportBatch
from django.contrib import messages
 
def reports_view(request):
    batches = Batch.objects.all().order_by("-created_at")
    batch_id = request.GET.get("batch")
    search_query = request.GET.get("search", "")  # search input
    severity_filter = request.GET.get("severity", "")  # optional filter
 
    if batch_id:
        selected_batch = get_object_or_404(Batch, id=batch_id)
    else:
        selected_batch = batches.first()
 
    defects = DefectLog.objects.filter(
        provision_job__provision__batch=selected_batch,
        provision_job__status="completed",
        provision_job__is_generated = False
    ).order_by("id") if selected_batch else DefectLog.objects.none()



    jobs = ProvisionJob.objects.filter(provision__batch = selected_batch, status = "completed", is_generated=False).order_by("id")
    jobs_with_defects = jobs.annotate(
        defect_count=Count('defect_logs')
    ).filter(defect_count__gt=0).order_by("id")
 

    severity_counts = {
    1: defects.filter(severity_level=1).count(),
    2: defects.filter(severity_level=2).count(),
    3: defects.filter(severity_level=3).count(),
    4: defects.filter(severity_level=4).count(),
        }
    severity_total = sum(severity_counts.values())

    quality_counts = {
    0: jobs.filter(document_rating=0).count(),
    1: jobs.filter(document_rating=1).count(),
    2: jobs.filter(document_rating=2).count(),
    3: jobs.filter(document_rating=3).count(),
}
    
    quality_total = sum(quality_counts.values())

    table1_data = []


   # Summary
    category_summary = {}
        
    jobs_with_defects_count = jobs_with_defects.count() if selected_batch else 0
    jobs_total = jobs.count() if selected_batch else 0
    provision_error_rate = (jobs_with_defects_count / jobs_total * 100) if jobs_total else 0

    for defect in defects:
        category_summary[defect.category] = category_summary.get(defect.category, 0) + 1
        if defect.screenshot:
            defect.screenshot_url = defect.get_absolute_url(request)
        else:
            defect.screenshot_url = ''

    greater_five_severity3_defects = defects.filter(severity_level=3, error_count__gt=5)
    greater_ten_severity4_defects = defects.filter(severity_level=4, error_count__gt=10)


    for defect in greater_five_severity3_defects:
        if defect.screenshot:
            defect.screenshot_url = defect.get_absolute_url(request)
        else:
            defect.screenshot_url = ""
    for defect in greater_ten_severity4_defects:
        if defect.screenshot:
            defect.screenshot_url = defect.get_absolute_url(request)
        else:
            defect.screenshot_url = ""
    # Apply search
    if search_query:
        defects = defects.filter(
            Q(issue_description__icontains=search_query) |
            Q(expected_outcome__icontains=search_query) |
            Q(actual_outcome__icontains=search_query) |
            Q(provision_job__provision__title__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(provision_job__enactment_assignment__enactment__title__icontains=search_query)
        )
 
    # Apply severity filter
    if severity_filter:
        defects = defects.filter(severity_level=severity_filter)
 
    # Pagination
    paginator = Paginator(defects, 10)  # 10 defects per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
 
 

        

    defect_categories = DefectCategory.objects.all()
        # Build defect options dynamically from DB
    defect_options = {}
    categories = DefectCategory.objects.prefetch_related("options").all()

    grand_total = 0

    for category in categories:
        options_list = [
            {
                "check_type": option.check_type,
                "severity_level": option.severity_level,
                "error_count": defects.filter(check_type=option.check_type).count()
            }
            for option in category.options.all()
        ]
        total_errors = sum(opt["error_count"] for opt in options_list)

        defect_options[category.name] = {
            "options": options_list,
            "total": total_errors,
        }

        grand_total += total_errors
        

    
    context = {
        "batches": batches,
        "selected_batch": selected_batch,
       "jobs":jobs,
        "defects": page_obj,  # paginated
        "total_defects": defects.count(),
        "category_summary": category_summary,
        "search_query": search_query,
        "severity_filter": severity_filter,

         "severity_counts": severity_counts,
        "severity_total": severity_total,
        "quality_counts": quality_counts,
        "quality_total": quality_total,
        "jobs_with_defects_count": jobs_with_defects_count,
        "jobs_total": jobs_total,
        "provision_error_rate": provision_error_rate,
        "active_page":"reports",
         "defect_categories": defect_categories,
    "defect_options": defect_options,
    "defect_options_grand_total":grand_total,
    "greater_five_severity3_defects":greater_five_severity3_defects,
    "greater_ten_severity4_defects":greater_ten_severity4_defects
    }

    return render(request, "reports/index.html", context)
 
 
 
def partial_excel_report(request):
    batch_id = request.GET.get("batch")
    batch = get_object_or_404(Batch, id=batch_id)
   
    jobs = ProvisionJob.objects.filter(
            provision__batch=batch,
            status="completed",
            is_generated=False
                    ).annotate(
                        defect_count=Count('defect_logs')
                    ).annotate(
                        review_outcome_text=Case(
                            When(defect_count__gt=0, then=Value("Defect Found")),
                            default=Value("No Defect Found"),
                            output_field=CharField()
                        )
                    )

    defects = DefectLog.objects.filter(
        provision_job__provision__batch=batch,
        provision_job__status="completed",
        provision_job__is_generated=False
    ).order_by("id")


    report_batch = ReportBatch.objects.create(created_by = request.user, batch = batch, batch_type = "partial")

    return generate_excel_response(batch, report_batch,jobs, defects, request)


def full_excel_report(request):
    batch_id = request.GET.get("batch")
    batch = get_object_or_404(Batch, id=batch_id)
    jobs = ProvisionJob.objects.filter(
            provision__batch=batch,
            status="completed"
    
                    ).annotate(
                        defect_count=Count('defect_logs')
                    ).annotate(
                        review_outcome_text=Case(
                            When(defect_count__gt=0, then=Value("Defect Found")),
                            default=Value("No Defect Found"),
                            output_field=CharField()
                        )
                    )

    defects = DefectLog.objects.filter(
        provision_job__provision__batch=batch,
        provision_job__status="completed",
        
    ).order_by("id")


    report_batch = ReportBatch.objects.create(created_by = request.user, batch = batch, batch_type = "full")


    return generate_excel_response(batch, report_batch,jobs, defects, request)


def generate_excel_response(batch,report_batch,jobs,defects, request=None):
    template_path = os.path.join(settings.BASE_DIR, "reports", "templates_excel", "output_template.xlsx")

    
    greater_five_severity3_defects = defects.filter(severity_level=3, error_count__gt=5)
    greater_ten_severity4_defects = defects.filter(severity_level=4, error_count__gt=10)
    
    # Load template (preserve formulas)
    wb = load_workbook(template_path, data_only=False)
    template_ws = wb["Summary"]

    # Access or create Defect Log sheet
    if "Defect Log" in wb.sheetnames:
        ws1 = wb["Defect Log"]
        # Clear existing data but keep header (assume header is in row 1)
        ws1.delete_rows(2, ws1.max_row)
    else:
        ws1 = wb.create_sheet("Defect Log")
        headers = [
            "Defect ID", "Enactment Citation", "Provision Ref(s)", "Version Date",
            "Category", "Check Type", "Issue Description", "Expected Outcome(BES)",
            "Actual Outcome(L+CP)", "Screenshot/Link", "Count per document",
            "Logged By", "Date Logged", "Comments",
        ]
        ws1.append(headers)

    # Access or create Enactments sheet
    if "Enactments" in wb.sheetnames:
        ws2 = wb["Enactments"]
        # Clear existing data but keep header (assume header is in row 1)
        ws2.delete_rows(2, ws2.max_row)
    else:
        ws2 = wb.create_sheet("Enactments")
        headers = [
            "Filename", "Enactment citation", "Provision", "Date \n (dd/mm/yyyy)",
            "Document Rating", "Review Outcome", "Remarks"
        ]
        ws2.append(headers)

    if "Error Type Highlights" in wb.sheetnames:
        ws4 = wb["Error Type Highlights"]
    


    table1_data = []
    table1_start_row = 6
    for defect in greater_five_severity3_defects:
        url = defect.get_absolute_url(request) if defect.screenshot and request else ""
        table1_data.append([
            defect.id,
            defect.provision_job.enactment_assignment.enactment.title,
            defect.provision_job.provision.title,
            defect.provision_job.date.strftime("%m/%d/%Y").lstrip("0").replace("/0", "/"),
            defect.category,
            defect.check_type,
            defect.issue_description,
            defect.expected_outcome,
            defect.actual_outcome,
            url,
            defect.severity_level,
         
            defect.error_count
        ])
    thin_border = Side(style="thin", color="000000")
    thick_border = Side(style="thick", color="000000")
    ws4.insert_rows(6,len(table1_data))

    table2_data = []
    for defect in greater_ten_severity4_defects:
        url = defect.get_absolute_url(request) if defect.screenshot and request else ""
        table2_data.append([
            defect.id,
            defect.provision_job.enactment_assignment.enactment.title,
            defect.provision_job.provision.title,
            defect.provision_job.date.strftime("%m/%d/%Y").lstrip("0").replace("/0", "/"),
            defect.category,
            defect.check_type,
            defect.issue_description,
            defect.expected_outcome,
            defect.actual_outcome,
            url,
             defect.severity_level,
         
            defect.error_count
        ])




    for i, row_data in enumerate(table1_data):
        for idx, j in enumerate(row_data):
            new_cell = ws4.cell(row=i+table1_start_row, column=idx+2, value=j)
            new_cell.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
            

    ws4.row_dimensions[table1_start_row+len(table1_data)+4].height = ws4.row_dimensions[5].height

    table2_data_start = len(table1_data) + 11
   
    for i, row_data in enumerate(table2_data):
        for idx, j in enumerate(row_data):
            new_cell = ws4.cell(row=i+table2_data_start, column=idx+2, value=j)
            new_cell.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

    for defect in defects:
        url = defect.get_absolute_url(request) if defect.screenshot and request else ''
        ws1.append([
            defect.id,
            defect.provision_job.enactment_assignment.enactment.title,
            defect.provision_job.provision.title,
            defect.provision_job.date.strftime("%d/%m/%Y"),
            defect.category,
            defect.check_type,
            defect.issue_description,
            defect.expected_outcome,
            defect.actual_outcome,
            "View Screenshot" if url else "",
            defect.severity_level,
            defect.error_count,
            defect.provision_job.user.email,
            defect.created_at.strftime("%d/%m/%Y"),
            defect.comments
        ])
        
        # get the last row index
        row_num = ws1.max_row

        # apply hyperlink and style if URL exists
        if url:
            cell = ws1.cell(row=row_num, column=10)   # column 10 = screenshot column
            cell.hyperlink = url
            cell.font = Font(color="0000FF", underline="single") 

    for job in jobs:
        review_outcome = "No Defect Found"
        if hasattr(job, "defect_logs") and job.defect_logs.exists():
            review_outcome = "Defect Found"
        elif hasattr(job, "review_outcome_text") and job.review_outcome_text:
            review_outcome = str(job.review_outcome_text)

    
        ws2.append([
            job.filename or "",
            job.provision.enactment.title if job.provision.enactment else "",
            job.provision.title if job.provision else "",
            job.date.strftime("%d/%m/%Y") if job.date else "",
            job.document_rating,
            review_outcome,
            job.remarks
        ])

        # set jobs generated
    for job in jobs:
        job.is_generated = True
        job.generation_date = timezone.now()
        job.save()

    report_batch.jobs.add(*jobs)




    # Generate filename
    batch_type = report_batch.batch_type.capitalize()
    batch_title = batch.name.replace(" ", "")
    date_str = datetime.now().strftime("%d%m%Y")
    filename = f"{batch_type}_{batch_title}_{date_str}.xlsx"

    # Create temp file and close immediately so openpyxl can write
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()  # now openpyxl can write to it

    # Save workbook to temp file for ReportBatch
    wb.save(tmp.name)

    # Save to ReportBatch file
    with open(tmp.name, "rb") as f:
        report_batch.file.save(filename, File(f), save=True)


    os.remove(tmp.name)

    # Save workbook to in-memory stream for download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def autofit_columns(ws):
    for col in ws.columns:
        column = col[0].column  # column number
        if 6 <= column <= 9:    # skip columns F to I
            continue
 
        max_length = 0
        column_letter = get_column_letter(column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

def add_outer_border(ws, cell_range, border_style="thin", color="000000"):
    """
    Apply an outer border around a range (like a rectangle) without overriding existing borders.
    
    :param ws: Worksheet
    :param cell_range: Excel range string (e.g., "B2:I3")
    :param border_style: Border style (default "thin")
    :param color: Border color (default black)
    """
    side = Side(border_style=border_style, color=color)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            
            # Get the current border of the cell
            current = cell.border

            # Only override sides if we are on the edge
            left   = side if col == min_col else current.left
            right  = side if col == max_col else current.right
            top    = side if row == min_row else current.top
            bottom = side if row == max_row else current.bottom

            # Reassign border while preserving non-edge sides
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def write_row_with_style(ws, row_idx, data, start_col=2, template_row=6):
    """
    Write row data to worksheet while preserving styles from a template row.
    """
    for j, value in enumerate(data, start=start_col):
        cell = ws.cell(row=row_idx, column=j, value=value)
        template_cell = ws.cell(row=template_row, column=j)
        if template_cell.has_style:
            cell._style = copy(template_cell._style)

from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from defects.models import DefectLog

@csrf_exempt
def edit_defect_log(request, defect_id):
    if request.method == "POST":
        defect = get_object_or_404(DefectLog, id=defect_id)

        # Update fields
        defect.category = request.POST.get("category", defect.category)
        defect.check_type = request.POST.get("check_type", defect.check_type)
        defect.severity_level = request.POST.get("severity_level", defect.severity_level)
        defect.issue_description = request.POST.get("issue_description", defect.issue_description)
        defect.expected_outcome = request.POST.get("expected_outcome", defect.expected_outcome)
        defect.actual_outcome = request.POST.get("actual_outcome", defect.actual_outcome)
        defect.comments = request.POST.get("comments", defect.comments)
        defect.error_count = request.POST.get("error_count", defect.error_count)

        if "screenshot" in request.FILES:
            defect.screenshot = request.FILES["screenshot"]

        defect.save()
        messages.success(request, "Defect log saved successfully")

        # Redirect to reports page or job detail if you want
        return redirect("reports_index")  # change "reports" to your desired route name

    return JsonResponse({"error": "Invalid request method."}, status=405)


def report_generation_detail(request, id):
    report_batch = get_object_or_404(ReportBatch, id=id)
    jobs_list = report_batch.jobs.all().order_by("id")  # optional ordering

    # Paginate: 10 jobs per page
    paginator = Paginator(jobs_list, 10)
    page_number = request.GET.get("page")
    jobs_page = paginator.get_page(page_number)

    context = {
        "report_batch": report_batch,
        "jobs_page": jobs_page,
    }
    return render(request, "reports/report_batch_detail.html", context)